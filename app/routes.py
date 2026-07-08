from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, send_file, session, g
from sqlalchemy import or_, func
import csv
import os
from io import StringIO, BytesIO
from unicodedata import normalize
from . import db
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from .models import Produto, Equipamento, TipoProduto, Local, Movimentacao, DashboardGrafico, ImportacaoPlanilha, Usuario

main = Blueprint("main", __name__)
_usuarios_iniciais_verificados = False


def _destino_seguro(destino):
    if destino and destino.startswith("/") and not destino.startswith("//"):
        return destino
    return url_for("main.dashboard")


def _garantir_usuario_admin():
    global _usuarios_iniciais_verificados
    if _usuarios_iniciais_verificados:
        return

    db.create_all()
    if Usuario.query.count() == 0:
        admin_email = os.environ.get("ADMIN_EMAIL", "admin@snpro.local").strip().lower()
        admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
        db.session.add(Usuario(
            nome="Administrador",
            email=admin_email,
            senha_hash=generate_password_hash(admin_password),
            perfil="Administrador",
            status="Ativo",
        ))
        db.session.commit()
    _usuarios_iniciais_verificados = True


def _usuario_eh_admin():
    return bool(getattr(g, "usuario_atual", None) and g.usuario_atual.perfil == "Administrador")


def _exigir_admin():
    if _usuario_eh_admin():
        return None
    flash("Apenas administradores podem acessar esta area.", "error")
    return redirect(url_for("main.dashboard"))


def _qtd_admins_ativos(excluir_id=None):
    query = Usuario.query.filter_by(perfil="Administrador", status="Ativo")
    if excluir_id:
        query = query.filter(Usuario.id != excluir_id)
    return query.count()


@main.before_app_request
def carregar_usuario_logado():
    if request.endpoint == "static":
        return None

    _garantir_usuario_admin()

    g.usuario_atual = None
    usuario_id = session.get("usuario_id")
    if usuario_id:
        usuario = Usuario.query.get(usuario_id)
        if usuario and usuario.status == "Ativo":
            g.usuario_atual = usuario
        else:
            session.clear()

    if request.endpoint in {"main.login"}:
        return None

    if not g.usuario_atual:
        proxima = request.full_path if request.method == "GET" else request.path
        return redirect(url_for("main.login", proxima=proxima))

    return None

EQUIPAMENTOS_PADRAO = [
    ("Periferico", "Itens como mouse, teclado, monitor, webcam e headset"),
    ("Desktop", "Computadores, gabinetes e equipamentos de mesa"),
    ("Notebook", "Notebooks, carregadores e componentes relacionados"),
    ("Servidor", "Servidores, pecas e componentes de infraestrutura"),
]

TIPOS_PRODUTO_PADRAO = [
    ("SSD", "Unidades de estado solido"),
    ("Processador", "CPUs e componentes de processamento"),
    ("Memória", "Memorias RAM e modulos relacionados"),
    ("Hard Disk", "Discos rigidos e armazenamentos magneticos"),
]
LOCAIS_PADRAO = [
    ("EP-Prateleira 3A", "Estoque Principal", "Prateleira principal para itens menores"),
    ("EP-Prateleira 3B", "Estoque Principal", "Prateleira intermediaria do estoque"),
    ("EP-Prateleira 3C", "Estoque Principal", "Prateleira superior do estoque"),
]


def _int_form(nome_campo, valor_padrao=0):
    valor = request.form.get(nome_campo)
    if valor in (None, ""):
        return valor_padrao
    return int(valor)


def _float_form(nome_campo, valor_padrao=0.0):
    valor = request.form.get(nome_campo)
    if valor in (None, ""):
        return valor_padrao
    return float(valor.replace(".", "").replace(",", "."))


def _normalizar_chave(valor):
    valor = normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode("ascii")
    return valor.strip().lower().replace(" ", "_").replace("-", "_")


def _impacto_estoque(tipo, quantidade):
    if tipo in ["Entrada", "Retorno"]:
        return quantidade
    if tipo in ["Saida", "Empréstimo", "Manutenção", "Descarte"]:
        return -quantidade
    return 0


def _registrar_movimentacao(produto, tipo, quantidade, valor_unitario, local, observacao=None, criado_em=None):
    if quantidade <= 0:
        raise ValueError("a quantidade precisa ser maior que zero")

    impacto = _impacto_estoque(tipo, quantidade)
    if impacto < 0 and produto.quantidade + impacto < 0:
        raise ValueError("quantidade indisponivel em estoque")

    if tipo not in ["Entrada", "Saida", "Transferência", "Empréstimo", "Retorno", "Manutenção", "Descarte"]:
        raise ValueError("tipo de movimentacao invalido")

    produto.quantidade += impacto
    if tipo == "Transferência" or local:
        produto.local = local or produto.local

    mov = Movimentacao(
        tipo=tipo,
        produto_id=produto.id,
        produto_nome=produto.nome,
        quantidade=quantidade,
        valor_unitario=valor_unitario,
        total=quantidade * valor_unitario,
        local=local or produto.local,
        observacao=observacao or None,
        criado_em=criado_em or datetime.now(),
    )
    db.session.add(mov)
    return mov


def _ler_linhas_planilha(arquivo):
    nome = (arquivo.filename or "").lower()
    if nome.endswith(".csv"):
        texto = arquivo.stream.read().decode("utf-8-sig", errors="replace")
        amostra = texto[:2048]
        try:
            dialecto = csv.Sniffer().sniff(amostra, delimiters=",;")
        except Exception:
            dialecto = csv.excel
            dialecto.delimiter = ";"
        leitor = csv.DictReader(StringIO(texto), dialect=dialecto)
        return list(leitor)

    if nome.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as erro:
            raise RuntimeError("para importar Excel, instale a dependencia openpyxl com: pip install -r requirements.txt") from erro
        wb = load_workbook(arquivo.stream, read_only=True, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            return []
        cabecalhos = [_normalizar_chave(c) for c in linhas[0]]
        dados = []
        for linha in linhas[1:]:
            if not any(c not in (None, "") for c in linha):
                continue
            dados.append({cabecalhos[i]: linha[i] if i < len(linha) else None for i in range(len(cabecalhos))})
        return dados

    raise ValueError("formato invalido. Envie um arquivo .csv ou .xlsx")



def _valor_vazio(valor):
    if valor is None:
        return True
    texto = str(valor).strip()
    return texto == "" or texto.lower() in {"nan", "none", "null", "nulo", "-", "#n/a", "n/a"}


def _texto_padrao(valor, padrao):
    return str(padrao if _valor_vazio(valor) else valor).strip()


def _sku_auto(indice):
    return f"AUTO-{datetime.now().strftime('%Y%m%d%H%M%S')}-{indice:04d}"

def _valor_linha(linha, *nomes, padrao=None):
    normalizada = {_normalizar_chave(k): v for k, v in linha.items() if k is not None}
    for nome in nomes:
        chave = _normalizar_chave(nome)
        if chave in normalizada and not _valor_vazio(normalizada[chave]):
            return normalizada[chave]
    return padrao


def _to_int(valor, padrao=0):
    if valor in (None, ""):
        return padrao
    try:
        return int(float(str(valor).replace(",", ".")))
    except Exception:
        return padrao


def _to_float(valor, padrao=0.0):
    if valor in (None, ""):
        return padrao
    txt = str(valor).strip().replace("R$", "").strip()
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return padrao


def _garantir_equipamentos_padrao():
    if Equipamento.query.count() == 0:
        for nome, descricao in EQUIPAMENTOS_PADRAO:
            db.session.add(Equipamento(nome=nome, descricao=descricao, status="Ativo"))
        db.session.commit()


def _nomes_equipamentos_ativos():
    _garantir_equipamentos_padrao()
    return [c.nome for c in Equipamento.query.filter_by(status="Ativo").order_by(Equipamento.nome.asc()).all()]


def _garantir_tipos_padrao():
    if TipoProduto.query.count() == 0:
        for nome, descricao in TIPOS_PRODUTO_PADRAO:
            db.session.add(TipoProduto(nome=nome, descricao=descricao, status="Ativo"))
        db.session.commit()


def _nomes_tipos_ativos():
    _garantir_tipos_padrao()
    return [t.nome for t in TipoProduto.query.filter_by(status="Ativo").order_by(TipoProduto.nome.asc()).all()]


def _garantir_locais_padrao():
    if Local.query.count() == 0:
        for nome, endereco, descricao in LOCAIS_PADRAO:
            db.session.add(Local(nome=nome, endereco=endereco, descricao=descricao, status="Ativo"))
        db.session.commit()


def _nomes_locais_ativos():
    _garantir_locais_padrao()
    return [l.nome for l in Local.query.filter_by(status="Ativo").order_by(Local.nome.asc()).all()]


DASHBOARD_FONTES = {
    "produtos": {
        "nome": "Produtos",
        "model": Produto,
        "colunas": {
            "equipamento": Produto.equipamento,
            "tipo_produto": Produto.tipo_produto,
            "local": Produto.local,
            "status": Produto.status,
            "status_estoque": None,
        },
        "metricas": {
            "contagem": ("Quantidade de registros", None),
            "soma_quantidade": ("Soma da quantidade", Produto.quantidade),
        },
        "filtros": {
            "sku": Produto.sku,
            "nome": Produto.nome,
            "equipamento": Produto.equipamento,
            "tipo_produto": Produto.tipo_produto,
            "local": Produto.local,
            "status": Produto.status,
            "quantidade": Produto.quantidade,
            "estoque_minimo": Produto.estoque_minimo,
        },
    },
    "movimentacoes": {
        "nome": "Movimentacoes",
        "model": Movimentacao,
        "colunas": {
            "tipo": Movimentacao.tipo,
            "produto_nome": Movimentacao.produto_nome,
            "local": Movimentacao.local,
            "data": None,
        },
        "metricas": {
            "contagem": ("Quantidade de registros", None),
            "soma_quantidade": ("Soma da quantidade", Movimentacao.quantidade),
            "soma_total": ("Soma da quantidade movimentada", Movimentacao.quantidade),
        },
        "filtros": {
            "tipo": Movimentacao.tipo,
            "produto_nome": Movimentacao.produto_nome,
            "local": Movimentacao.local,
            "quantidade": Movimentacao.quantidade,
        },
    },
}

DASHBOARD_COLUNAS_LABEL = {
    "equipamento": "Equipamento",
    "tipo_produto": "Tipo de Produto",
    "local": "Local",
    "status": "Status",
    "status_estoque": "Status do Estoque",
    "tipo": "Tipo",
    "produto_nome": "Produto",
    "data": "Data",
    "sku": "SKU",
    "nome": "Nome",
    "quantidade": "Quantidade",
    "estoque_minimo": "Estoque Minimo",
}


def _aplicar_filtro_dashboard(query, config, grafico):
    coluna_nome = (grafico.filtro_coluna or "").strip()
    valor = (grafico.filtro_valor or "").strip()
    operador = grafico.filtro_operador or "contem"
    coluna = config["filtros"].get(coluna_nome)

    if not coluna_nome or coluna is None or valor == "":
        return query

    try:
        if operador == "igual":
            return query.filter(coluna == valor)
        if operador == "diferente":
            return query.filter(coluna != valor)
        if operador == "maior":
            return query.filter(coluna > _to_float(valor, 0))
        if operador == "menor":
            return query.filter(coluna < _to_float(valor, 0))
        return query.filter(coluna.ilike(f"%{valor}%"))
    except Exception:
        return query


def _dados_grafico(grafico):
    config = DASHBOARD_FONTES.get(grafico.fonte, DASHBOARD_FONTES["produtos"])
    model = config["model"]
    query = _aplicar_filtro_dashboard(model.query, config, grafico)
    coluna_nome = grafico.coluna_grupo
    metrica = grafico.metrica

    dados = []

    if grafico.fonte == "produtos" and coluna_nome == "status_estoque":
        produtos = query.all()
        grupos = {}
        for produto in produtos:
            chave = produto.status_estoque or "Sem valor"
            if metrica == "soma_quantidade":
                valor = produto.quantidade or 0
            else:
                valor = 1
            grupos.setdefault(chave, []).append(valor)
        for chave, valores in grupos.items():
            total = sum(valores)
            dados.append({"label": chave, "value": round(total, 2)})
        return dados

    if grafico.fonte == "movimentacoes" and coluna_nome == "data":
        linhas = query.order_by(Movimentacao.criado_em.asc()).all()
        grupos = {}
        for mov in linhas:
            chave = mov.criado_em.strftime("%d/%m/%Y") if mov.criado_em else "Sem data"
            if metrica == "soma_quantidade":
                valor = mov.quantidade or 0
            elif metrica == "soma_total":
                valor = mov.quantidade or 0
            else:
                valor = 1
            grupos.setdefault(chave, []).append(valor)
        for chave, valores in grupos.items():
            total = sum(valores)
            dados.append({"label": chave, "value": round(total, 2)})
        return dados[-20:]

    coluna = config["colunas"].get(coluna_nome)
    if coluna is None:
        return []

    if metrica == "contagem":
        consulta = query.with_entities(coluna, func.count(model.id)).group_by(coluna).order_by(func.count(model.id).desc())
    elif metrica.startswith("media"):
        campo = config["metricas"].get(metrica, (None, None))[1]
        consulta = query.with_entities(coluna, func.avg(campo)).group_by(coluna).order_by(func.avg(campo).desc()) if campo is not None else []
    else:
        campo = config["metricas"].get(metrica, (None, None))[1]
        consulta = query.with_entities(coluna, func.sum(campo)).group_by(coluna).order_by(func.sum(campo).desc()) if campo is not None else []

    for label, value in consulta:
        dados.append({"label": str(label or "Sem valor"), "value": round(float(value or 0), 2)})
    return dados[:20]


def _dashboard_resumo():
    total_produtos = Produto.query.count()
    total_itens = db.session.query(func.coalesce(func.sum(Produto.quantidade), 0)).scalar() or 0
    baixo_estoque = Produto.query.filter(Produto.quantidade <= Produto.estoque_minimo).count()
    return {
        "total_produtos": total_produtos,
        "total_itens": int(total_itens),
        "baixo_estoque": baixo_estoque,
    }


def _limpar_graficos_antigos_de_valor():
    metricas_removidas = {"soma_valor_estoque", "media_ticket", "media_valor_unitario"}
    alterou = False
    for grafico in DashboardGrafico.query.all():
        if grafico.metrica in metricas_removidas or "valor" in (grafico.titulo or "").lower() or "ticket" in (grafico.titulo or "").lower():
            db.session.delete(grafico)
            alterou = True
    if alterou:
        db.session.commit()


def _garantir_graficos_padrao():
    _limpar_graficos_antigos_de_valor()
    if DashboardGrafico.query.count() == 0:
        padroes = [
            DashboardGrafico(titulo="Produtos por equipamento", fonte="produtos", tipo_grafico="barra", coluna_grupo="equipamento", metrica="contagem", criado_em=datetime.now()),
            DashboardGrafico(titulo="Produtos por local", fonte="produtos", tipo_grafico="barra", coluna_grupo="local", metrica="soma_quantidade", criado_em=datetime.now()),
            DashboardGrafico(titulo="Movimentacoes por tipo", fonte="movimentacoes", tipo_grafico="pizza", coluna_grupo="tipo", metrica="contagem", criado_em=datetime.now()),
        ]
        db.session.add_all(padroes)
        db.session.commit()


@main.route("/login", methods=["GET", "POST"])
def login():
    if getattr(g, "usuario_atual", None) and request.method == "GET":
        return redirect(_destino_seguro(request.args.get("proxima")))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        senha = request.form.get("senha", "")
        usuario = Usuario.query.filter_by(email=email).first()

        if usuario and usuario.status == "Ativo" and check_password_hash(usuario.senha_hash, senha):
            session.clear()
            session["usuario_id"] = usuario.id
            flash("Login realizado com sucesso.", "success")
            return redirect(_destino_seguro(request.args.get("proxima")))

        flash("E-mail ou senha invalidos.", "error")

    return render_template("login.html")


@main.route("/logout")
def logout():
    session.clear()
    flash("Voce saiu do sistema.", "success")
    return redirect(url_for("main.login"))


@main.route("/usuarios")
def usuarios():
    bloqueio = _exigir_admin()
    if bloqueio:
        return bloqueio

    usuarios_lista = Usuario.query.order_by(Usuario.nome.asc()).all()
    return render_template("usuarios.html", active_page="usuarios", usuarios=usuarios_lista)


@main.route("/usuarios/novo", methods=["POST"])
def novo_usuario():
    bloqueio = _exigir_admin()
    if bloqueio:
        return bloqueio

    try:
        senha = request.form.get("senha", "")
        if len(senha) < 6:
            raise ValueError("a senha precisa ter pelo menos 6 caracteres")

        usuario = Usuario(
            nome=request.form["nome"].strip(),
            email=request.form["email"].strip().lower(),
            senha_hash=generate_password_hash(senha),
            perfil=request.form.get("perfil") or "Operador",
            status=request.form.get("status") or "Ativo",
        )
        db.session.add(usuario)
        db.session.commit()
        flash("Usuario cadastrado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao cadastrar usuario: {erro}", "error")

    return redirect(url_for("main.usuarios"))


@main.route("/usuarios/editar/<int:usuario_id>", methods=["POST"])
def editar_usuario(usuario_id):
    bloqueio = _exigir_admin()
    if bloqueio:
        return bloqueio

    usuario = Usuario.query.get_or_404(usuario_id)

    try:
        perfil_novo = request.form.get("perfil") or "Operador"
        status_novo = request.form.get("status") or "Ativo"
        removeria_admin = usuario.perfil == "Administrador" and (perfil_novo != "Administrador" or status_novo != "Ativo")
        if removeria_admin and _qtd_admins_ativos(excluir_id=usuario.id) == 0:
            raise ValueError("nao e possivel remover o ultimo administrador ativo")

        usuario.nome = request.form["nome"].strip()
        usuario.email = request.form["email"].strip().lower()
        usuario.perfil = perfil_novo
        usuario.status = status_novo

        senha = request.form.get("senha", "")
        if senha:
            if len(senha) < 6:
                raise ValueError("a senha precisa ter pelo menos 6 caracteres")
            usuario.senha_hash = generate_password_hash(senha)

        db.session.commit()
        flash("Usuario atualizado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar usuario: {erro}", "error")

    return redirect(url_for("main.usuarios"))


@main.route("/usuarios/excluir/<int:usuario_id>", methods=["POST"])
def excluir_usuario(usuario_id):
    bloqueio = _exigir_admin()
    if bloqueio:
        return bloqueio

    usuario = Usuario.query.get_or_404(usuario_id)

    if g.usuario_atual and usuario.id == g.usuario_atual.id:
        flash("Voce nao pode excluir seu proprio usuario.", "error")
        return redirect(url_for("main.usuarios"))

    if usuario.perfil == "Administrador" and usuario.status == "Ativo" and _qtd_admins_ativos(excluir_id=usuario.id) == 0:
        flash("Nao e possivel excluir o ultimo administrador ativo.", "error")
        return redirect(url_for("main.usuarios"))

    db.session.delete(usuario)
    db.session.commit()
    flash("Usuario excluido com sucesso.", "success")
    return redirect(url_for("main.usuarios"))


@main.route("/dashboard")
def dashboard():
    _garantir_equipamentos_padrao()
    _garantir_tipos_padrao()
    _garantir_locais_padrao()
    _garantir_graficos_padrao()

    graficos = DashboardGrafico.query.order_by(DashboardGrafico.criado_em.asc()).all()
    graficos_render = []
    for grafico in graficos:
        config = DASHBOARD_FONTES.get(grafico.fonte, DASHBOARD_FONTES["produtos"])
        metrica_label = config["metricas"].get(grafico.metrica, (grafico.metrica, None))[0]
        graficos_render.append({
            "obj": grafico,
            "dados": _dados_grafico(grafico),
            "metrica_label": metrica_label,
            "coluna_label": DASHBOARD_COLUNAS_LABEL.get(grafico.coluna_grupo, grafico.coluna_grupo),
        })

    return render_template(
        "dashboard.html",
        active_page="dashboard",
        resumo=_dashboard_resumo(),
        graficos=graficos_render,
        fontes=DASHBOARD_FONTES,
        colunas_label=DASHBOARD_COLUNAS_LABEL,
    )


@main.route("/dashboard/graficos/novo", methods=["POST"])
def novo_grafico_dashboard():
    try:
        grafico = DashboardGrafico(
            titulo=request.form.get("titulo") or "Grafico personalizado",
            fonte=request.form.get("fonte") or "produtos",
            tipo_grafico=request.form.get("tipo_grafico") or "barra",
            coluna_grupo=request.form.get("coluna_grupo") or "equipamento",
            metrica=request.form.get("metrica") or "contagem",
            filtro_coluna=request.form.get("filtro_coluna") or None,
            filtro_operador=request.form.get("filtro_operador") or None,
            filtro_valor=request.form.get("filtro_valor") or None,
            criado_em=datetime.now(),
        )
        db.session.add(grafico)
        db.session.commit()
        flash("Grafico adicionado ao dashboard.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao criar grafico: {erro}", "error")
    return redirect(url_for("main.dashboard"))


@main.route("/dashboard/graficos/excluir/<int:grafico_id>", methods=["POST"])
def excluir_grafico_dashboard(grafico_id):
    grafico = DashboardGrafico.query.get_or_404(grafico_id)
    db.session.delete(grafico)
    db.session.commit()
    flash("Grafico removido do dashboard.", "success")
    return redirect(url_for("main.dashboard"))


@main.route("/")
def index():
    return redirect(url_for("main.dashboard"))


@main.route("/produtos")
def produtos():
    busca = request.args.get("busca", "").strip()
    equipamento = request.args.get("equipamento", "").strip()
    local = request.args.get("local", "").strip()
    status = request.args.get("status", "").strip()

    query = Produto.query

    if busca:
        query = query.filter(or_(Produto.nome.ilike(f"%{busca}%"), Produto.sku.ilike(f"%{busca}%")))
    if equipamento:
        query = query.filter(Produto.equipamento == equipamento)
    if local:
        query = query.filter(Produto.local == local)
    if status:
        query = query.filter(Produto.status == status)

    produtos_lista = query.order_by(Produto.nome.asc()).all()

    return render_template(
        "produtos.html",
        active_page="produtos",
        produtos=produtos_lista,
        equipamentos=_nomes_equipamentos_ativos(),
        tipos_produto=_nomes_tipos_ativos(),
        locais=_nomes_locais_ativos(),
        busca=busca,
        equipamento=equipamento,
        local=local,
        status=status,
    )


@main.route("/produtos/novo", methods=["POST"])
def novo_produto():
    try:
        produto = Produto(
            sku=request.form["sku"].strip(),
            nome=request.form["nome"].strip(),
            equipamento=request.form["equipamento"],
            tipo_produto=request.form.get("tipo_produto") or None,
            local=request.form["local"],
            quantidade=_int_form("quantidade", 1),
            estoque_minimo=_int_form("estoque_minimo", 1),
            ticket_medio=0.0,
            status=request.form.get("status") or "Ativo",
            descricao=request.form.get("descricao") or None,
        )
        db.session.add(produto)
        db.session.commit()
        flash("Produto cadastrado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao cadastrar produto: {erro}", "error")

    return redirect(url_for("main.produtos"))


@main.route("/produtos/editar/<int:produto_id>", methods=["POST"])
def editar_produto(produto_id):
    produto = Produto.query.get_or_404(produto_id)

    try:
        produto.sku = request.form["sku"].strip()
        produto.nome = request.form["nome"].strip()
        produto.equipamento = request.form["equipamento"]
        produto.tipo_produto = request.form.get("tipo_produto") or None
        produto.local = request.form["local"]
        produto.quantidade = _int_form("quantidade", 1)
        produto.estoque_minimo = _int_form("estoque_minimo", 1)
        produto.ticket_medio = 0.0
        produto.status = request.form.get("status") or "Ativo"
        produto.descricao = request.form.get("descricao") or None

        db.session.commit()
        flash("Produto atualizado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar produto: {erro}", "error")

    return redirect(url_for("main.produtos"))


@main.route("/produtos/excluir/<int:produto_id>", methods=["POST"])
def excluir_produto(produto_id):
    produto = Produto.query.get_or_404(produto_id)
    db.session.delete(produto)
    db.session.commit()
    flash("Produto excluido com sucesso.", "success")
    return redirect(url_for("main.produtos"))


@main.route("/equipamentos")
def equipamentos():
    _garantir_equipamentos_padrao()
    equipamentos_lista = Equipamento.query.order_by(Equipamento.nome.asc()).all()
    return render_template("equipamentos.html", active_page="equipamentos", equipamentos=equipamentos_lista)


@main.route("/equipamentos/nova", methods=["POST"])
def nova_equipamento():
    try:
        equipamento = Equipamento(
            nome=request.form["nome"].strip(),
            descricao=request.form.get("descricao") or None,
            status=request.form.get("status") or "Ativo",
        )
        db.session.add(equipamento)
        db.session.commit()
        flash("Equipamento cadastrado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao cadastrar equipamento: {erro}", "error")
    return redirect(url_for("main.equipamentos"))


@main.route("/equipamentos/editar/<int:equipamento_id>", methods=["POST"])
def editar_equipamento(equipamento_id):
    equipamento = Equipamento.query.get_or_404(equipamento_id)
    nome_antigo = equipamento.nome

    try:
        equipamento.nome = request.form["nome"].strip()
        equipamento.descricao = request.form.get("descricao") or None
        equipamento.status = request.form.get("status") or "Ativo"

        if nome_antigo != equipamento.nome:
            Produto.query.filter_by(equipamento=nome_antigo).update({"equipamento": equipamento.nome})

        db.session.commit()
        flash("Equipamento atualizado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar equipamento: {erro}", "error")
    return redirect(url_for("main.equipamentos"))


@main.route("/equipamentos/excluir/<int:equipamento_id>", methods=["POST"])
def excluir_equipamento(equipamento_id):
    equipamento = Equipamento.query.get_or_404(equipamento_id)

    produto_vinculado = Produto.query.filter_by(equipamento=equipamento.nome).first()
    if produto_vinculado:
        flash("Nao foi possivel excluir: existem produtos usando este equipamento.", "error")
        return redirect(url_for("main.equipamentos"))

    db.session.delete(equipamento)
    db.session.commit()
    flash("Equipamento excluido com sucesso.", "success")
    return redirect(url_for("main.equipamentos"))


@main.route("/tipos-produto")
def tipos_produto():
    _garantir_tipos_padrao()
    tipos_lista = TipoProduto.query.order_by(TipoProduto.nome.asc()).all()
    return render_template("tipos_produto.html", active_page="tipos_produto", tipos=tipos_lista)


@main.route("/tipos-produto/novo", methods=["POST"])
def novo_tipo_produto():
    try:
        tipo = TipoProduto(
            nome=request.form["nome"].strip(),
            descricao=request.form.get("descricao") or None,
            status=request.form.get("status") or "Ativo",
        )
        db.session.add(tipo)
        db.session.commit()
        flash("Tipo de produto cadastrado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao cadastrar tipo de produto: {erro}", "error")
    return redirect(url_for("main.tipos_produto"))


@main.route("/tipos-produto/editar/<int:tipo_id>", methods=["POST"])
def editar_tipo_produto(tipo_id):
    tipo = TipoProduto.query.get_or_404(tipo_id)
    nome_antigo = tipo.nome

    try:
        tipo.nome = request.form["nome"].strip()
        tipo.descricao = request.form.get("descricao") or None
        tipo.status = request.form.get("status") or "Ativo"

        if nome_antigo != tipo.nome:
            Produto.query.filter_by(tipo_produto=nome_antigo).update({"tipo_produto": tipo.nome})

        db.session.commit()
        flash("Tipo de produto atualizado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar tipo de produto: {erro}", "error")
    return redirect(url_for("main.tipos_produto"))


@main.route("/tipos-produto/excluir/<int:tipo_id>", methods=["POST"])
def excluir_tipo_produto(tipo_id):
    tipo = TipoProduto.query.get_or_404(tipo_id)

    produto_vinculado = Produto.query.filter_by(tipo_produto=tipo.nome).first()
    if produto_vinculado:
        flash("Nao foi possivel excluir: existem produtos usando este tipo de produto.", "error")
        return redirect(url_for("main.tipos_produto"))

    db.session.delete(tipo)
    db.session.commit()
    flash("Tipo de produto excluido com sucesso.", "success")
    return redirect(url_for("main.tipos_produto"))


@main.route("/locais")
def locais():
    _garantir_locais_padrao()
    locais_lista = Local.query.order_by(Local.nome.asc()).all()
    return render_template("locais.html", active_page="locais", locais=locais_lista)


@main.route("/locais/novo", methods=["POST"])
def novo_local():
    try:
        local = Local(
            nome=request.form["nome"].strip(),
            endereco=request.form.get("endereco") or None,
            descricao=request.form.get("descricao") or None,
            status=request.form.get("status") or "Ativo",
        )
        db.session.add(local)
        db.session.commit()
        flash("Local cadastrado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao cadastrar local: {erro}", "error")
    return redirect(url_for("main.locais"))


@main.route("/locais/editar/<int:local_id>", methods=["POST"])
def editar_local(local_id):
    local = Local.query.get_or_404(local_id)
    nome_antigo = local.nome

    try:
        local.nome = request.form["nome"].strip()
        local.endereco = request.form.get("endereco") or None
        local.descricao = request.form.get("descricao") or None
        local.status = request.form.get("status") or "Ativo"

        if nome_antigo != local.nome:
            Produto.query.filter_by(local=nome_antigo).update({"local": local.nome})

        db.session.commit()
        flash("Local atualizado com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar local: {erro}", "error")
    return redirect(url_for("main.locais"))


@main.route("/locais/excluir/<int:local_id>", methods=["POST"])
def excluir_local(local_id):
    local = Local.query.get_or_404(local_id)

    produto_vinculado = Produto.query.filter_by(local=local.nome).first()
    if produto_vinculado:
        flash("Nao foi possivel excluir: existem produtos usando este local.", "error")
        return redirect(url_for("main.locais"))

    db.session.delete(local)
    db.session.commit()
    flash("Local excluido com sucesso.", "success")
    return redirect(url_for("main.locais"))


@main.route("/movimentacoes")
def movimentacoes():
    busca = request.args.get("busca", "").strip()
    tipo = request.args.get("tipo", "").strip()
    produto_id = request.args.get("produto_id", "").strip()

    query = Movimentacao.query

    if busca:
        query = query.filter(Movimentacao.produto_nome.ilike(f"%{busca}%"))
    if tipo:
        query = query.filter(Movimentacao.tipo == tipo)
    if produto_id:
        query = query.filter(Movimentacao.produto_id == int(produto_id))

    movimentacoes_lista = query.order_by(Movimentacao.criado_em.desc()).all()
    produtos_lista = Produto.query.order_by(Produto.nome.asc()).all()

    return render_template(
        "movimentacoes.html",
        active_page="movimentacoes",
        movimentacoes=movimentacoes_lista,
        produtos=produtos_lista,
        locais=_nomes_locais_ativos(),
        busca=busca,
        tipo=tipo,
        produto_id=produto_id,
    )


@main.route("/movimentacoes/nova", methods=["POST"])
def nova_movimentacao():
    try:
        produto = Produto.query.get_or_404(int(request.form["produto_id"]))
        tipo = request.form["tipo"].strip()
        quantidade = _int_form("quantidade", 1)
        valor_unitario = 0.0
        local = request.form.get("local") or produto.local

        _registrar_movimentacao(
            produto=produto,
            tipo=tipo,
            quantidade=quantidade,
            valor_unitario=valor_unitario,
            local=local,
            observacao=request.form.get("observacao"),
        )
        db.session.commit()
        flash("Movimentacao registrada com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao registrar movimentacao: {erro}", "error")

    return redirect(url_for("main.movimentacoes"))


@main.route("/movimentacoes/editar/<int:movimentacao_id>", methods=["POST"])
def editar_movimentacao(movimentacao_id):
    movimentacao = Movimentacao.query.get_or_404(movimentacao_id)

    try:
        produto_antigo = Produto.query.get(movimentacao.produto_id)
        if produto_antigo:
            produto_antigo.quantidade -= _impacto_estoque(movimentacao.tipo, movimentacao.quantidade)
            if produto_antigo.quantidade < 0:
                produto_antigo.quantidade = 0

        produto = Produto.query.get_or_404(int(request.form["produto_id"]))
        tipo = request.form["tipo"].strip()
        quantidade = _int_form("quantidade", 1)
        valor_unitario = 0.0
        local = request.form.get("local") or produto.local

        impacto = _impacto_estoque(tipo, quantidade)
        if impacto < 0 and produto.quantidade + impacto < 0:
            raise ValueError("quantidade indisponivel em estoque para a correcao")

        produto.quantidade += impacto
        produto.local = local
        produto.ticket_medio = 0.0

        movimentacao.tipo = tipo
        movimentacao.produto_id = produto.id
        movimentacao.produto_nome = produto.nome
        movimentacao.quantidade = quantidade
        movimentacao.valor_unitario = valor_unitario
        movimentacao.total = quantidade * valor_unitario
        movimentacao.local = local
        movimentacao.observacao = request.form.get("observacao") or None

        data_mov = request.form.get("criado_em")
        if data_mov:
            movimentacao.criado_em = datetime.strptime(data_mov, "%Y-%m-%dT%H:%M")

        db.session.commit()
        flash("Movimentacao atualizada com sucesso.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao atualizar movimentacao: {erro}", "error")

    return redirect(url_for("main.movimentacoes"))


@main.route("/movimentacoes/excluir/<int:movimentacao_id>", methods=["POST"])
def excluir_movimentacao(movimentacao_id):
    movimentacao = Movimentacao.query.get_or_404(movimentacao_id)
    db.session.delete(movimentacao)
    db.session.commit()
    flash("Movimentacao excluida com sucesso.", "success")
    return redirect(url_for("main.movimentacoes"))



def _tipo_relatorio_valido(tipo):
    if tipo in {"produtos", "baixo_estoque", "movimentacoes"}:
        return tipo
    return "produtos"


def _conteudo_relatorio(tipo):
    tipo = _tipo_relatorio_valido(tipo)

    if tipo == "movimentacoes":
        cabecalhos = ["Tipo", "Produto", "Quantidade", "Local", "Data", "Observacao"]
        linhas = [
            [
                mov.tipo,
                mov.produto_nome,
                mov.quantidade,
                mov.local,
                mov.criado_em.strftime("%d/%m/%Y %H:%M") if mov.criado_em else "",
                mov.observacao or "",
            ]
            for mov in Movimentacao.query.order_by(Movimentacao.criado_em.desc()).all()
        ]
        return "Movimentacoes", "relatorio_movimentacoes", cabecalhos, linhas

    lista = Produto.query
    titulo = "Produtos Cadastrados"
    nome_arquivo = "relatorio_produtos"
    if tipo == "baixo_estoque":
        lista = lista.filter(Produto.quantidade <= Produto.estoque_minimo)
        titulo = "Produtos com Estoque Baixo"
        nome_arquivo = "relatorio_baixo_estoque"

    cabecalhos = ["Codigo", "Nome", "Equipamento", "Tipo", "Local", "Quantidade", "Estoque Minimo", "Status", "Descricao"]
    linhas = [
        [
            produto.sku,
            produto.nome,
            produto.equipamento,
            produto.tipo_produto or "",
            produto.local,
            produto.quantidade,
            produto.estoque_minimo,
            produto.status,
            produto.descricao or "",
        ]
        for produto in lista.order_by(Produto.nome.asc()).all()
    ]
    return titulo, nome_arquivo, cabecalhos, linhas


@main.route("/relatorios")
def relatorios():
    tipo = _tipo_relatorio_valido(request.args.get("tipo", "produtos").strip())

    produtos_lista = Produto.query.order_by(Produto.nome.asc()).all()
    movimentacoes_lista = Movimentacao.query.order_by(Movimentacao.criado_em.desc()).all()
    baixo_estoque = Produto.query.filter(Produto.quantidade <= Produto.estoque_minimo).order_by(Produto.nome.asc()).all()
    importacoes_count = ImportacaoPlanilha.query.count()

    return render_template(
        "relatorios.html",
        active_page="relatorios",
        tipo=tipo,
        produtos=produtos_lista,
        movimentacoes=movimentacoes_lista,
        baixo_estoque=baixo_estoque,
        importacoes_count=importacoes_count,
    )


@main.route("/relatorios/exportar")
def exportar_relatorio():
    tipo = _tipo_relatorio_valido(request.args.get("tipo", "produtos"))
    saida = StringIO()
    writer = csv.writer(saida, delimiter=";")
    _, nome_arquivo, cabecalhos, linhas = _conteudo_relatorio(tipo)

    writer.writerow(cabecalhos)
    writer.writerows(linhas)

    conteudo = "\ufeff" + saida.getvalue()
    return Response(
        conteudo,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.csv"},
    )


@main.route("/relatorios/exportar-excel")
def exportar_relatorio_excel():
    tipo = _tipo_relatorio_valido(request.args.get("tipo", "produtos"))
    titulo, nome_arquivo, cabecalhos, linhas = _conteudo_relatorio(tipo)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as erro:
        flash(f"Erro ao exportar Excel: dependencia openpyxl indisponivel ({erro}).", "error")
        return redirect(url_for("main.relatorios", tipo=tipo))

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ws.append(cabecalhos)
    for linha in linhas:
        ws.append(linha)

    header_fill = PatternFill("solid", fgColor="D9E2EF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for coluna in ws.columns:
        letra = get_column_letter(coluna[0].column)
        maior = max(len(str(cell.value or "")) for cell in coluna)
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 45)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"{nome_arquivo}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main.route("/relatorios/importar", methods=["POST"])
def importar_planilha():
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo CSV ou Excel para importar.", "error")
        return redirect(url_for("main.relatorios"))

    try:
        linhas = _ler_linhas_planilha(arquivo)
        importados = 0
        atualizados = 0
        skus_importados = []

        for indice, linha in enumerate(linhas, start=1):
            sku = _texto_padrao(_valor_linha(linha, "sku", "codigo", "código", "cod", "Código"), _sku_auto(indice))
            nome = _texto_padrao(_valor_linha(linha, "nome", "produto", "nome_produto"), f"Produto importado {indice}")

            equipamento = _texto_padrao(_valor_linha(linha, "equipamento"), "Sem Equipamento")
            tipo_produto = _texto_padrao(_valor_linha(linha, "tipo", "tipo_produto"), "Sem Tipo")
            local = _texto_padrao(_valor_linha(linha, "local", "locais", "armazenamento"), "Estoque Principal")
            quantidade = _to_int(_valor_linha(linha, "quantidade", "qtd"), 0)
            estoque_minimo = _to_int(_valor_linha(linha, "estoque_minimo", "estoque mínimo", "limite_estoque_baixo"), 0)
            status = _texto_padrao(_valor_linha(linha, "status"), "Ativo")
            descricao = None if _valor_vazio(_valor_linha(linha, "descricao", "descrição")) else _valor_linha(linha, "descricao", "descrição")

            if not Equipamento.query.filter_by(nome=equipamento).first():
                db.session.add(Equipamento(nome=equipamento, descricao="Importado da planilha", status="Ativo"))
            if tipo_produto and not TipoProduto.query.filter_by(nome=tipo_produto).first():
                db.session.add(TipoProduto(nome=tipo_produto, descricao="Importado da planilha", status="Ativo"))
            if not Local.query.filter_by(nome=local).first():
                db.session.add(Local(nome=local, endereco="Importado da planilha", descricao="Importado da planilha", status="Ativo"))

            produto = Produto.query.filter_by(sku=sku).first()
            skus_importados.append(sku)

            if produto:
                produto.nome = nome
                produto.equipamento = equipamento
                produto.tipo_produto = tipo_produto
                produto.local = local
                produto.quantidade = quantidade
                produto.estoque_minimo = estoque_minimo
                produto.ticket_medio = 0.0
                produto.status = status if status in ["Ativo", "Inativo"] else "Ativo"
                produto.descricao = descricao
                atualizados += 1
            else:
                produto = Produto(
                    sku=sku,
                    nome=nome,
                    equipamento=equipamento,
                    tipo_produto=tipo_produto,
                    local=local,
                    quantidade=quantidade,
                    estoque_minimo=estoque_minimo,
                    ticket_medio=0.0,
                    status=status if status in ["Ativo", "Inativo"] else "Ativo",
                    descricao=descricao,
                )
                db.session.add(produto)
                importados += 1

        registro_importacao = ImportacaoPlanilha(
            nome_arquivo=arquivo.filename,
            produtos_skus="\n".join(sorted(set(skus_importados))),
            total_novos=importados,
            total_atualizados=atualizados,
            criado_em=datetime.now(),
        )
        db.session.add(registro_importacao)

        db.session.commit()
        flash(f"Importacao concluida: {importados} produtos novos e {atualizados} atualizados.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao importar planilha: {erro}", "error")

    return redirect(url_for("main.relatorios"))


@main.route("/relatorios/importacoes/apagar", methods=["POST"])
def apagar_dados_importados():
    try:
        importacoes = ImportacaoPlanilha.query.all()
        skus = set()
        for importacao in importacoes:
            for sku in (importacao.produtos_skus or "").splitlines():
                sku = sku.strip()
                if sku:
                    skus.add(sku)

        if not skus:
            equipamentos_importados = [c.nome for c in Equipamento.query.filter_by(descricao="Importado da planilha").all()]
            tipos_importados = [t.nome for t in TipoProduto.query.filter_by(descricao="Importado da planilha").all()]
            locais_importados = [l.nome for l in Local.query.filter_by(descricao="Importado da planilha").all()]
            produtos = Produto.query.filter(
                or_(
                    Produto.sku.like("AUTO-%"),
                    Produto.equipamento.in_(equipamentos_importados or ["__sem_equipamento_importado__"]),
                    Produto.tipo_produto.in_(tipos_importados or ["__sem_tipo_importado__"]),
                    Produto.local.in_(locais_importados or ["__sem_local_importado__"]),
                )
            ).all()
        else:
            produtos = Produto.query.filter(Produto.sku.in_(skus)).all()
        produto_ids = [produto.id for produto in produtos]

        if produto_ids:
            Movimentacao.query.filter(Movimentacao.produto_id.in_(produto_ids)).delete(synchronize_session=False)

        apagados = 0
        for produto in produtos:
            db.session.delete(produto)
            apagados += 1

        ImportacaoPlanilha.query.delete()

        for equipamento in Equipamento.query.filter_by(descricao="Importado da planilha").all():
            if not Produto.query.filter_by(equipamento=equipamento.nome).first():
                db.session.delete(equipamento)
        for tipo in TipoProduto.query.filter_by(descricao="Importado da planilha").all():
            if not Produto.query.filter_by(tipo_produto=tipo.nome).first():
                db.session.delete(tipo)
        for local in Local.query.filter_by(descricao="Importado da planilha").all():
            if not Produto.query.filter_by(local=local.nome).first():
                db.session.delete(local)

        db.session.commit()
        flash(f"Dados importados apagados com sucesso: {apagados} produtos removidos.", "success")
    except Exception as erro:
        db.session.rollback()
        flash(f"Erro ao apagar dados importados: {erro}", "error")

    return redirect(url_for("main.relatorios"))


@main.route("/seed")
def seed():
    _garantir_equipamentos_padrao()
    _garantir_tipos_padrao()
    _garantir_locais_padrao()
    if Produto.query.count() == 0:
        exemplos = [
            ("MPC20152", "Memoria Smart PC2 1RX8 1GB"),
            ("MPC20151", "Memoria Smart PC2 1RX8 1GB"),
            ("MPC22433", "Memoria Smart PC2 1RX4 1GB"),
            ("MPC20142", "Memoria Markvision PC2 1GB"),
            ("MPC20150", "Memoria Kingston PC2 1GB"),
        ]
        for sku, nome in exemplos:
            produto = Produto(
                sku=sku,
                nome=nome,
                equipamento="Periferico",
                tipo_produto="Memoria",
                local="EP-Prateleira 3C",
                quantidade=1,
                estoque_minimo=1,
                ticket_medio=0.0,
                status="Ativo"
            )
            db.session.add(produto)
            db.session.flush()
            db.session.add(Movimentacao(
                tipo="Entrada",
                produto_id=produto.id,
                produto_nome=produto.nome,
                quantidade=1,
                valor_unitario=0.0,
                total=0.0,
                local="EP-Prateleira 3C",
                criado_em=datetime.now(),
            ))
        db.session.commit()
        flash("Produtos e movimentacoes de exemplo adicionados.", "success")
    else:
        flash("O banco ja possui produtos.", "error")
    return redirect(url_for("main.produtos"))
